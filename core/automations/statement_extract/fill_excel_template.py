from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font
import pandas as pd
import os

def fill_excel_template(df: pd.DataFrame, output_path: str, nome: str):
    # Carrega o workbook e a primeira planilha
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, '../../templates/Planilha_Extratos_TEMPLATE.xlsx')
    
    wb = load_workbook(template_path)
    ws = wb.active

    # Insere o nome no cabeçalho (células A2:C2 mescladas)
    ws.merge_cells('A2:C2')
    ws['A2'] = nome

    # Preparar cópia de estilo da linha modelo (linha 4)
    template_row_idx = 4
    styles = {}
    for col_idx in range(1, 4):  # colunas A, B, C
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}{template_row_idx}"]
        styles[col_letter] = {
            'font': cell.font.copy(),
            'border': cell.border.copy(),
            'fill': cell.fill.copy(),
            'alignment': cell.alignment.copy(),
            'number_format': cell.number_format,
            'protection': cell.protection.copy(),
        }
        
    df = df.reset_index(drop=True)
    print(f"DataFrame filtrado:\n{df}")

    start_row = template_row_idx
    n_rows = len(df)

    # Insere linhas em branco para os dados
    ws.insert_rows(start_row, amount=n_rows)

    # Preenche as linhas com dados e aplica estilos
    for i, row in df.iterrows():
        excel_row = start_row + i
        
        for col_idx, field in enumerate(['Data', 'Historico', 'Valor'], start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{excel_row}"]
            # Aplica valor
            if field == 'Valor':
                valor_str = str(row[field]).replace('.', '').replace(',', '.')
                try:
                    cell.value = float(valor_str)
                except ValueError:
                    cell.value = 0.0
            else:
                cell.value = row[field]
            # Aplica cópia do estilo
            style = styles[col_letter]
            cell.font = style['font']
            cell.border = style['border']
            if col_letter == 'C':
                cell.fill = style['fill']
            else:
                cell.fill = PatternFill(fill_type=None)
            cell.alignment = style['alignment']
            cell.number_format = style['number_format']
            cell.protection = style['protection']
            if col_letter == 'B':
                cell.font = Font(**{**style['font'].__dict__, "bold": True})
            elif col_letter == 'A':
                cell.font = Font(**{**style['font'].__dict__, "bold": False})
            

    # Reaplica as mesclagens no footer após os dados
    footer_start = start_row + n_rows  # linha onde o footer começa agora

    for merge in list(ws.merged_cells.ranges):
        if merge.min_row >= footer_start:
            ws.unmerge_cells(str(merge))

    # Linha 1 do footer: normal (nada a fazer)
    # Linha 2 do footer: A:C mesclado
    ws.merge_cells(f"A{footer_start+1}:C{footer_start+1}")

    # Linha 3 do footer: B:C mesclado
    ws.merge_cells(f"B{footer_start+2}:C{footer_start+2}")
    ws[f"B{footer_start+2}"] = df["Historico"][df.index[-1]]  # último histórico

    # Linha 4 do footer: B:C mesclado
    ws.merge_cells(f"B{footer_start+3}:C{footer_start+3}")

    # Linha 5 do footer: B:C mesclado
    ws.merge_cells(f"B{footer_start+4}:C{footer_start+4}")
    
    df['Valor'] = (
        df['Valor']
        .astype(str)
        .str.replace('.', '', regex=False)   # Remove pontos (milhares)
        .str.replace(',', '.', regex=False)  # Substitui vírgula por ponto (decimal)
        .astype(float)                       # Converte para float
    )
    total = df["Valor"].sum()
    # total = df["Valor"].replace(",", ".", regex=True).astype(float).sum()

    ws[f"C{footer_start}"] = total  # ou use footer_start+1 se for na linha 1 do footer
    ws[f"B{footer_start+3}"] = f"=C{footer_start}"
    ws[f"B{footer_start+4}"] = f"=C{footer_start}*2"

    # Salva o arquivo de saída
    wb.save(output_path)
